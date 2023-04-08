from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
wb = load_workbook ('D:\\test excel\\test1.xlsx')
ws = wb['Sheet1']
sheet = wb.active
a = sheet.max_row
b=[]
i=0
while i < a:
    i = i+1
    print(ws.cell (i,2).value)
    b.insert(i, ws.cell (i,2).value)

wbtest = load_workbook ('D:\\test excel\\test2.xlsx')
ws = wbtest['Sheet1']
sheet = wb.active
i = 0
while i < a:
    ws.cell (i+1,1).value = b[i]
    i = i+1
wbtest.save ('C:\\Users\\elan\\Desktop\\python\\test.xlsx')